"""
데이터 분석 및 내보내기 모듈

[파일 역할]
데이터베이스에 저장된 주식 데이터를 분석하고 다양한 형식으로 내보냅니다.

[주요 기능]
- Excel/CSV 내보내기
- 통계 분석 (수익률, 변동성, 샤프 비율 등)
- 자동 리포트 생성
- Grafana/Power BI 연동 지원

[사용 방법]
from data_analyzer import DataAnalyzer
from database import StockDatabase

db = StockDatabase()
analyzer = DataAnalyzer(db)
analyzer.export_to_excel('005930', start_date, end_date, 'output.xlsx')
"""

from datetime import datetime, timedelta
from typing import Optional, Dict, List
from pathlib import Path
import math

from utils.logger import log


class DataAnalyzer:
    """
    데이터 분석 및 내보내기 클래스
    
    저장된 주식 데이터를 분석하고 다양한 포맷으로 내보냅니다.
    """
    
    def __init__(self, database):
        """
        Args:
            database: StockDatabase 인스턴스
        """
        self.database = database
        
        if not database or not database.enabled:
            log.warning("데이터베이스가 비활성화되어 분석 기능을 사용할 수 없습니다.")
            self.enabled = False
        else:
            self.enabled = True
            log.info("DataAnalyzer 초기화 완료")
    
    def export_to_csv(
        self,
        stock_code: str,
        start_date: datetime,
        end_date: datetime,
        output_path: str
    ) -> bool:
        """
        CSV 파일로 내보내기
        
        Args:
            stock_code: 종목 코드
            start_date: 시작 날짜
            end_date: 종료 날짜
            output_path: 출력 파일 경로
            
        Returns:
            성공 여부
        """
        if not self.enabled:
            return False
        
        try:
            # 데이터 조회
            candles = self.database.get_candles(stock_code, start_date, end_date)
            
            if not candles:
                log.warning(f"CSV 내보내기 실패: 데이터 없음 ({stock_code})")
                return False
            
            # CSV 생성
            import csv
            
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=[
                    'timestamp', 'stock_code', 'open', 'high', 'low', 'close', 'volume'
                ])
                writer.writeheader()
                
                for candle in candles:
                    writer.writerow({
                        'timestamp': candle['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                        'stock_code': candle['stock_code'],
                        'open': candle['open'],
                        'high': candle['high'],
                        'low': candle['low'],
                        'close': candle['close'],
                        'volume': candle['volume']
                    })
            
            log.success(f"CSV 내보내기 완료: {output_path} ({len(candles)}개)")
            return True
            
        except Exception as e:
            log.error(f"CSV 내보내기 오류: {e}")
            return False
    
    def export_to_excel(
        self,
        stock_code: str,
        start_date: datetime,
        end_date: datetime,
        output_path: str
    ) -> bool:
        """
        Excel 파일로 내보내기
        
        Args:
            stock_code: 종목 코드
            start_date: 시작 날짜
            end_date: 종료 날짜
            output_path: 출력 파일 경로
            
        Returns:
            성공 여부
        """
        if not self.enabled:
            return False
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            # 데이터 조회
            candles = self.database.get_candles(stock_code, start_date, end_date)
            
            if not candles:
                log.warning(f"Excel 내보내기 실패: 데이터 없음 ({stock_code})")
                return False
            
            # 통계 계산
            stats = self.get_statistics(stock_code, start_date, end_date)
            
            # 엑셀 워크북 생성
            wb = Workbook()
            
            # 시트 1: 1분봉 데이터
            ws_candles = wb.active
            ws_candles.title = "1분봉 데이터"
            
            # 헤더
            headers = ['날짜/시간', '종목코드', '시가', '고가', '저가', '종가', '거래량']
            ws_candles.append(headers)
            
            # 헤더 스타일
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            
            for col in range(1, len(headers) + 1):
                cell = ws_candles.cell(1, col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            # 데이터 추가
            for candle in candles:
                ws_candles.append([
                    candle['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                    candle['stock_code'],
                    candle['open'],
                    candle['high'],
                    candle['low'],
                    candle['close'],
                    candle['volume']
                ])
            
            # 열 너비 자동 조정
            for col in range(1, len(headers) + 1):
                ws_candles.column_dimensions[get_column_letter(col)].width = 15
            
            # 시트 2: 통계
            if stats:
                ws_stats = wb.create_sheet("통계")
                
                ws_stats.append(['항목', '값'])
                ws_stats.append(['종목 코드', stock_code])
                ws_stats.append(['기간', f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}"])
                ws_stats.append([])
                ws_stats.append(['1분봉 개수', stats['candle_count']])
                ws_stats.append(['최저가', f"{stats['min_price']:,.0f}원"])
                ws_stats.append(['최고가', f"{stats['max_price']:,.0f}원"])
                ws_stats.append(['평균가', f"{stats['avg_price']:,.0f}원"])
                ws_stats.append(['총 거래량', f"{stats['total_volume']:,}"])
                ws_stats.append([])
                ws_stats.append(['변동성', f"{stats['volatility']:.2f}%"])
                ws_stats.append(['일평균 수익률', f"{stats['avg_daily_return']:.2f}%"])
                ws_stats.append(['최대 상승', f"{stats['max_gain']:.2f}%"])
                ws_stats.append(['최대 하락', f"{stats['max_loss']:.2f}%"])
                
                # 스타일
                for row in range(1, ws_stats.max_row + 1):
                    ws_stats.cell(row, 1).font = Font(bold=True)
                    ws_stats.cell(row, 1).alignment = Alignment(horizontal="left")
                    ws_stats.cell(row, 2).alignment = Alignment(horizontal="right")
                
                ws_stats.column_dimensions['A'].width = 20
                ws_stats.column_dimensions['B'].width = 25
            
            # 파일 저장
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            
            log.success(f"Excel 내보내기 완료: {output_path} ({len(candles)}개)")
            return True
            
        except ImportError:
            log.error("openpyxl이 설치되지 않았습니다. pip install openpyxl")
            return False
        except Exception as e:
            log.error(f"Excel 내보내기 오류: {e}")
            return False
    
    def get_statistics(
        self,
        stock_code: str,
        start_date: datetime,
        end_date: datetime
    ) -> Optional[Dict]:
        """
        통계 분석
        
        Args:
            stock_code: 종목 코드
            start_date: 시작 날짜
            end_date: 종료 날짜
            
        Returns:
            통계 정보 딕셔너리
        """
        if not self.enabled:
            return None
        
        try:
            # 데이터 조회
            candles = self.database.get_candles(stock_code, start_date, end_date)
            
            if not candles or len(candles) < 2:
                return None
            
            # 기본 통계
            prices = [c['close'] for c in candles]
            volumes = [c['volume'] for c in candles]
            
            min_price = min(prices)
            max_price = max(prices)
            avg_price = sum(prices) / len(prices)
            total_volume = sum(volumes)
            
            # 수익률 계산
            returns = []
            for i in range(1, len(prices)):
                ret = (prices[i] - prices[i-1]) / prices[i-1] * 100
                returns.append(ret)
            
            # 변동성 (표준편차)
            if returns:
                avg_return = sum(returns) / len(returns)
                variance = sum((r - avg_return) ** 2 for r in returns) / len(returns)
                volatility = math.sqrt(variance)
                
                max_gain = max(returns) if returns else 0
                max_loss = min(returns) if returns else 0
            else:
                avg_return = 0
                volatility = 0
                max_gain = 0
                max_loss = 0
            
            # 전체 기간 수익률
            total_return = (prices[-1] - prices[0]) / prices[0] * 100 if prices else 0
            
            # 일평균 수익률 (1분봉 → 일봉 변환)
            days = (end_date - start_date).days or 1
            avg_daily_return = total_return / days
            
            # 샤프 비율 (단순화: 무위험 수익률 0 가정)
            sharpe_ratio = (avg_return / volatility) if volatility > 0 else 0
            
            return {
                'stock_code': stock_code,
                'candle_count': len(candles),
                'min_price': min_price,
                'max_price': max_price,
                'avg_price': avg_price,
                'total_volume': total_volume,
                'total_return': total_return,
                'avg_daily_return': avg_daily_return,
                'volatility': volatility,
                'max_gain': max_gain,
                'max_loss': max_loss,
                'sharpe_ratio': sharpe_ratio,
                'start_date': start_date,
                'end_date': end_date
            }
            
        except Exception as e:
            log.error(f"통계 분석 오류 ({stock_code}): {e}")
            return None
    
    def generate_report(
        self,
        stock_code: str,
        start_date: datetime,
        end_date: datetime,
        output_path: str
    ) -> bool:
        """
        HTML 리포트 생성
        
        Args:
            stock_code: 종목 코드
            start_date: 시작 날짜
            end_date: 종료 날짜
            output_path: 출력 파일 경로
            
        Returns:
            성공 여부
        """
        if not self.enabled:
            return False
        
        try:
            # 통계 계산
            stats = self.get_statistics(stock_code, start_date, end_date)
            
            if not stats:
                log.warning(f"리포트 생성 실패: 데이터 없음 ({stock_code})")
                return False
            
            # HTML 생성
            html = f"""
<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{stock_code} 분석 리포트</title>
    <style>
        body {{
            font-family: 'Malgun Gothic', sans-serif;
            max-width: 800px;
            margin: 50px auto;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            background-color: white;
            padding: 30px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #333;
            border-bottom: 3px solid #4472C4;
            padding-bottom: 10px;
        }}
        .stats {{
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 20px;
            margin-top: 30px;
        }}
        .stat-box {{
            background: #f8f9fa;
            padding: 15px;
            border-radius: 5px;
            border-left: 4px solid #4472C4;
        }}
        .stat-label {{
            font-size: 14px;
            color: #666;
            margin-bottom: 5px;
        }}
        .stat-value {{
            font-size: 24px;
            font-weight: bold;
            color: #333;
        }}
        .positive {{ color: #d9534f; }}
        .negative {{ color: #5cb85c; }}
        .footer {{
            margin-top: 30px;
            padding-top: 20px;
            border-top: 1px solid #ddd;
            text-align: center;
            color: #999;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>📊 {stock_code} 분석 리포트</h1>
        
        <p><strong>분석 기간:</strong> {stats['start_date'].strftime('%Y-%m-%d')} ~ {stats['end_date'].strftime('%Y-%m-%d')}</p>
        <p><strong>생성 시간:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        
        <div class="stats">
            <div class="stat-box">
                <div class="stat-label">1분봉 개수</div>
                <div class="stat-value">{stats['candle_count']:,}개</div>
            </div>
            
            <div class="stat-box">
                <div class="stat-label">평균가</div>
                <div class="stat-value">{stats['avg_price']:,.0f}원</div>
            </div>
            
            <div class="stat-box">
                <div class="stat-label">최저가</div>
                <div class="stat-value">{stats['min_price']:,.0f}원</div>
            </div>
            
            <div class="stat-box">
                <div class="stat-label">최고가</div>
                <div class="stat-value">{stats['max_price']:,.0f}원</div>
            </div>
            
            <div class="stat-box">
                <div class="stat-label">총 수익률</div>
                <div class="stat-value {'positive' if stats['total_return'] >= 0 else 'negative'}">
                    {stats['total_return']:+.2f}%
                </div>
            </div>
            
            <div class="stat-box">
                <div class="stat-label">일평균 수익률</div>
                <div class="stat-value {'positive' if stats['avg_daily_return'] >= 0 else 'negative'}">
                    {stats['avg_daily_return']:+.2f}%
                </div>
            </div>
            
            <div class="stat-box">
                <div class="stat-label">변동성</div>
                <div class="stat-value">{stats['volatility']:.2f}%</div>
            </div>
            
            <div class="stat-box">
                <div class="stat-label">샤프 비율</div>
                <div class="stat-value">{stats['sharpe_ratio']:.2f}</div>
            </div>
            
            <div class="stat-box">
                <div class="stat-label">최대 상승</div>
                <div class="stat-value positive">{stats['max_gain']:+.2f}%</div>
            </div>
            
            <div class="stat-box">
                <div class="stat-label">최대 하락</div>
                <div class="stat-value negative">{stats['max_loss']:+.2f}%</div>
            </div>
        </div>
        
        <div class="footer">
            CleonAI 자동매매 프로그램 | 데이터 분석 리포트
        </div>
    </div>
</body>
</html>
"""
            
            # 파일 저장
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html)
            
            log.success(f"HTML 리포트 생성 완료: {output_path}")
            return True
            
        except Exception as e:
            log.error(f"리포트 생성 오류: {e}")
            return False
    
    def print_statistics(
        self,
        stock_code: str,
        start_date: datetime,
        end_date: datetime
    ):
        """
        통계를 콘솔에 출력
        
        Args:
            stock_code: 종목 코드
            start_date: 시작 날짜
            end_date: 종료 날짜
        """
        stats = self.get_statistics(stock_code, start_date, end_date)
        
        if not stats:
            print(f"❌ {stock_code} 통계 없음")
            return
        
        print("=" * 70)
        print(f"📊 {stock_code} 통계 분석")
        print("=" * 70)
        print(f"기간: {start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}")
        print(f"\n[기본 정보]")
        print(f"  1분봉 개수: {stats['candle_count']:,}개")
        print(f"  최저가: {stats['min_price']:,.0f}원")
        print(f"  최고가: {stats['max_price']:,.0f}원")
        print(f"  평균가: {stats['avg_price']:,.0f}원")
        print(f"  총 거래량: {stats['total_volume']:,}")
        print(f"\n[수익률]")
        print(f"  총 수익률: {stats['total_return']:+.2f}%")
        print(f"  일평균 수익률: {stats['avg_daily_return']:+.2f}%")
        print(f"\n[리스크]")
        print(f"  변동성: {stats['volatility']:.2f}%")
        print(f"  최대 상승: {stats['max_gain']:+.2f}%")
        print(f"  최대 하락: {stats['max_loss']:+.2f}%")
        print(f"  샤프 비율: {stats['sharpe_ratio']:.2f}")
        print("=" * 70)


if __name__ == "__main__":
    # 테스트 코드
    from database import StockDatabase
    
    print("=" * 70)
    print("DataAnalyzer 테스트")
    print("=" * 70)
    
    # 테스트용 데이터베이스
    db = StockDatabase("data/test_stocks.duckdb", "data/test_parquet")
    
    if not db.enabled:
        print("DuckDB가 설치되지 않았습니다.")
        exit(1)
    
    # 분석기 생성
    analyzer = DataAnalyzer(db)
    
    # 테스트 데이터 생성 (실제로는 trading_engine에서 수집)
    print("\n1. 테스트 데이터 생성...")
    test_candles = []
    base_time = datetime.now().replace(hour=9, minute=0, second=0, microsecond=0)
    base_price = 75000
    
    for i in range(100):
        import random
        price = base_price + random.randint(-1000, 1000)
        test_candles.append({
            'stock_code': '005930',
            'timestamp': base_time + timedelta(minutes=i),
            'open': price,
            'high': price + random.randint(0, 500),
            'low': price - random.randint(0, 500),
            'close': price + random.randint(-200, 200),
            'volume': random.randint(500000, 1500000)
        })
    
    db.save_candles_batch(test_candles)
    print(f"   {len(test_candles)}개 1분봉 저장 완료")
    
    # 기간 설정
    start_date = base_time
    end_date = base_time + timedelta(hours=2)
    
    # 통계 출력
    print("\n2. 통계 분석:")
    analyzer.print_statistics('005930', start_date, end_date)
    
    # CSV 내보내기
    print("\n3. CSV 내보내기...")
    csv_path = "data/output/005930_analysis.csv"
    if analyzer.export_to_csv('005930', start_date, end_date, csv_path):
        print(f"   ✅ {csv_path}")
    
    # Excel 내보내기
    print("\n4. Excel 내보내기...")
    excel_path = "data/output/005930_analysis.xlsx"
    if analyzer.export_to_excel('005930', start_date, end_date, excel_path):
        print(f"   ✅ {excel_path}")
    
    # HTML 리포트
    print("\n5. HTML 리포트 생성...")
    html_path = "data/output/005930_report.html"
    if analyzer.generate_report('005930', start_date, end_date, html_path):
        print(f"   ✅ {html_path}")
    
    print("\n" + "=" * 70)
    print("테스트 완료!")
    print("=" * 70)

